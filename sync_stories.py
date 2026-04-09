"""
sync_stories.py

Reads story text from moments-metadata.xlsx and updates the data-story
attributes in moments.html for each matching image filename.
"""

import re
import openpyxl


EXCEL_PATH = "/Users/royezuz/Desktop/roy-website/assets/moments-metadata.xlsx"
HTML_PATH = "/Users/royezuz/Desktop/roy-website/moments.html"


def load_stories_from_excel(path):
    """Return a dict mapping filename (lowercase) -> story text."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Find column indices from the header row
    headers = [cell.value for cell in ws[1]]
    try:
        filename_col = headers.index("Filename")
        story_col = headers.index("Story")
    except ValueError as e:
        raise RuntimeError(f"Expected column not found in Excel headers: {e}")

    stories = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        filename = row[filename_col]
        story = row[story_col]

        if not filename:
            continue
        if not story or str(story).strip() == "":
            continue

        stories[str(filename).strip().lower()] = str(story).strip()

    return stories


def escape_for_attr(text):
    """Escape characters that would break an HTML attribute value in double quotes."""
    # Replace literal double quotes with the HTML entity
    return text.replace('"', "&quot;")


def update_html_stories(html_path, stories):
    """
    For each filename in `stories`, find the matching <img> tag in the HTML
    and replace its data-story attribute value.

    Returns (updated_html, updated_count, already_current, not_found_in_html, attr_missing).
    """
    with open(html_path, "r", encoding="utf-8") as f:
        html = f.read()

    updated = 0
    already_current = []
    not_found_in_html = []
    attr_missing = []

    for filename_lower, story_text in stories.items():
        # Build a pattern that finds any <img … > that has this filename at the
        # end of its src attribute (case-insensitive on the filename portion).
        # We match the full tag so we can surgically replace just data-story.
        src_pattern = re.compile(
            r'(<img\b[^>]*?\bsrc=["\']images/moments/' + re.escape(filename_lower) + r'["\'][^>]*?>)',
            re.IGNORECASE | re.DOTALL,
        )

        match = src_pattern.search(html)
        if not match:
            not_found_in_html.append(filename_lower)
            continue

        original_tag = match.group(1)

        # Check whether data-story attribute exists at all
        if 'data-story="' not in original_tag:
            attr_missing.append(filename_lower)
            continue

        escaped_story = escape_for_attr(story_text)

        # Replace the data-story value (handles both empty and non-empty existing values)
        new_tag = re.sub(
            r'data-story="[^"]*"',
            f'data-story="{escaped_story}"',
            original_tag,
            count=1,
        )

        if new_tag == original_tag:
            # The value was already identical — nothing to do
            already_current.append(filename_lower)
            continue

        html = html[:match.start(1)] + new_tag + html[match.end(1):]
        updated += 1

    return html, updated, already_current, not_found_in_html, attr_missing


def main():
    print("Loading stories from Excel...")
    stories = load_stories_from_excel(EXCEL_PATH)
    print(f"  Found {len(stories)} non-empty stories in Excel.")

    print("Updating HTML...")
    new_html, updated, already_current, not_found_in_html, attr_missing = update_html_stories(
        HTML_PATH, stories
    )

    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(new_html)

    print(f"\nDone.")
    print(f"  Updated  : {updated} image(s) — story written to HTML")
    print(f"  Current  : {len(already_current)} image(s) — HTML already matched Excel, no change needed")

    if not_found_in_html:
        print(f"\nWarning — {len(not_found_in_html)} filename(s) from Excel not found in HTML:")
        for name in not_found_in_html:
            print(f"  - {name}")

    if attr_missing:
        print(f"\nWarning — {len(attr_missing)} image(s) found in HTML but have no data-story attribute:")
        for name in attr_missing:
            print(f"  - {name}")

    if already_current:
        print(f"\nAlready up to date:")
        for name in already_current:
            print(f"  - {name}")


if __name__ == "__main__":
    main()
